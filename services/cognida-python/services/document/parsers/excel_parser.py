"""Excel 文档解析器。"""

from typing import Any

from .base import BaseParser, ParseResult


class ExcelParser(BaseParser):
    """Excel (xlsx/xls) 文档解析器。"""

    @property
    def supported_formats(self) -> list[str]:
        return ["xlsx", "xls"]

    async def parse(
        self,
        source: str | bytes,
        include_metadata: bool = True,
        extract_tables: bool = True,
        extract_images: bool = False,
        **kwargs: Any,
    ) -> ParseResult:
        """解析 Excel 文档。

        Args:
            source: 文件路径或内容
            include_metadata: 是否包含元数据
            extract_tables: 是否提取表格
            extract_images: 是否提取图片
            **kwargs: 其他选项
                - sheet_name: 工作表名称（默认第一个）

        Returns:
            解析结果
        """
        try:
            import openpyxl
            from io import BytesIO

            if isinstance(source, str):
                workbook = openpyxl.load_workbook(source)
            else:
                workbook = openpyxl.load_workbook(BytesIO(source))

            sheet_name = kwargs.get("sheet_name")
            if sheet_name:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active

            # 提取文本
            text_parts = []
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                text_parts.append(row_text)
            text = "\n".join(text_parts)

            # 构建元数据
            metadata = {}
            if include_metadata:
                metadata = {
                    "format": "xlsx",
                    "sheet_count": len(workbook.worksheets),
                    "sheet_name": sheet.title,
                    "row_count": sheet.max_row or 0,
                    "column_count": sheet.max_column or 0,
                    "char_count": len(text),
                }

            # 提取表格（Excel 本身就是表格）
            tables = None
            if extract_tables:
                tables = self._extract_tables(sheet)

            return ParseResult(
                success=True,
                text=text,
                metadata=metadata,
                tables=tables,
            )

        except Exception as e:
            return ParseResult(
                success=False,
                text="",
                metadata={},
                error=str(e),
            )

    def _extract_tables(self, sheet: Any) -> list[dict[str, Any]]:
        """提取表格。

        Args:
            sheet: 工作表对象

        Returns:
            表格列表
        """
        rows = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            rows.append(cells)

        markdown = self._table_to_markdown(rows)

        return [{
            "table_index": 0,
            "rows": rows,
            "markdown": markdown,
        }]

    def _table_to_markdown(self, rows: list[list[str]]) -> str:
        """将表格转换为 Markdown 格式。

        Args:
            rows: 表格行数据

        Returns:
            Markdown 字符串
        """
        if not rows:
            return ""

        lines = []
        for i, row in enumerate(rows):
            cells = [cell.replace("|", "\\|").strip() for cell in row]
            lines.append("| " + " | ".join(cells) + " |")
            if i == 0:
                lines.append("|" + "|".join(["---"] * len(cells)) + "|")

        return "\n".join(lines)
